import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Design Tokens (harmonized color scheme matching the app)
PRIMARY_BLUE = "1961A5"  # EndoAI main theme color
DARK_BLUE = "0F396B"     # Headers
TITLE_FONT = Font(name="Outfit", size=16, bold=True, color=PRIMARY_BLUE)
HEADER_FONT = Font(name="Outfit", size=11, bold=True, color="FFFFFF")

# Status Fills
FILL_PASSED = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # Soft Green
FILL_FAILED = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # Soft Red
FILL_BLOCKED = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid") # Soft Orange
FILL_SKIPPED = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid") # Soft Gray

# Fills for Sheet Headers
HEADER_FILL = PatternFill(start_color=PRIMARY_BLUE, end_color=PRIMARY_BLUE, fill_type="solid")
ACCENT_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_THIN = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

def apply_auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if '\n' in val_str:
                max_len = max(max_len, max(len(s) for s in val_str.split('\n')))
            else:
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(min(max_len + 3, 50), 10)

def generate_appium_reports(test_cases, duration_sec):
    # Ensure directories exist
    os.makedirs("Test_Results", exist_ok=True)
    os.makedirs("Test_Results/Excel", exist_ok=True)
    os.makedirs("Test_Results/Markdown", exist_ok=True)
    
    total = len(test_cases)
    passed = sum(1 for tc in test_cases if tc["Status"] == "Passed")
    failed = sum(1 for tc in test_cases if tc["Status"] == "Failed")
    skipped = sum(1 for tc in test_cases if tc["Status"] == "Skipped")
    blocked = sum(1 for tc in test_cases if tc["Status"] == "Blocked")
    pass_percent = round((passed / total) * 100, 2) if total > 0 else 0.0
    
    # 1. Generate Excel Report
    wb = openpyxl.Workbook()
    
    # Active tab is Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    
    # Summary Sheet Design
    ws_sum.append(["EndoAI Mobile Appium E2E Test Suite Summary Report"])
    ws_sum.cell(row=1, column=1).font = TITLE_FONT
    ws_sum.append([])
    ws_sum.append(["Metric", "Value"])
    
    # Setup Metrics Rows
    metrics = [
        ("Total Test Cases", total),
        ("Passed Tests", passed),
        ("Failed Tests", failed),
        ("Skipped Tests", skipped),
        ("Blocked Tests", blocked),
        ("Pass Percentage", f"{pass_percent}%"),
        ("Execution Duration", f"{duration_sec:.2f} seconds"),
        ("Execution Status", "PASS" if pass_percent >= 95.0 else "FAIL")
    ]
    
    for metric, value in metrics:
        ws_sum.append([metric, value])
        
    # Style Summary Headers
    ws_sum.cell(row=3, column=1).font = HEADER_FONT
    ws_sum.cell(row=3, column=1).fill = HEADER_FILL
    ws_sum.cell(row=3, column=1).alignment = ALIGN_CENTER
    ws_sum.cell(row=3, column=2).font = HEADER_FONT
    ws_sum.cell(row=3, column=2).fill = HEADER_FILL
    ws_sum.cell(row=3, column=2).alignment = ALIGN_CENTER
    
    # Apply grid styles to Summary
    for row in range(3, 12):
        ws_sum.cell(row=row, column=1).border = BORDER_THIN
        ws_sum.cell(row=row, column=2).border = BORDER_THIN
        
        # Color highlight for Execution Status
        if row == 11:
            val_cell = ws_sum.cell(row=row, column=2)
            if pass_percent >= 95.0:
                val_cell.fill = FILL_PASSED
                val_cell.font = Font(name="Outfit", bold=True, color="155724")
            else:
                val_cell.fill = FILL_FAILED
                val_cell.font = Font(name="Outfit", bold=True, color="721C24")
                
    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 25
    
    # 2. Details Sheet
    ws_det = wb.create_sheet(title="Details")
    headers_det = [
        "Test ID", "Module", "Category", "Test Scenario", "Preconditions", 
        "Test Steps", "Test Data", "Expected Result", "Priority", 
        "Test Type", "Automation Status", "Execution Time (sec)", "Status", "Failure Reason"
    ]
    ws_det.append(headers_det)
    
    # Format headers
    for idx, col_name in enumerate(headers_det):
        cell = ws_det.cell(row=1, column=idx+1)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
        
    # Populate Details Rows
    for tc in test_cases:
        row_values = [
            tc["Test Case ID"],
            tc["Module"],
            tc["Category"],
            tc["Test Scenario"],
            tc["Preconditions"],
            tc["Test Steps"],
            tc["Test Data"],
            tc["Expected Result"],
            tc["Priority"],
            tc["Test Type"],
            tc["Automation Status"],
            tc["Execution Time"],
            tc["Status"],
            tc["Failure Reason"]
        ]
        ws_det.append(row_values)
        
    # Format detail records
    for row_idx in range(2, len(test_cases) + 2):
        for col_idx in range(1, len(headers_det) + 1):
            cell = ws_det.cell(row=row_idx, column=col_idx)
            cell.border = BORDER_THIN
            
            # Alignments
            if col_idx in [1, 9, 10, 11, 12, 13]:  # ID, Priority, Type, Status, time
                cell.alignment = ALIGN_CENTER
            else:
                cell.alignment = ALIGN_LEFT
                
            # Status colors
            if col_idx == 13: # Status column
                status_val = cell.value
                if status_val == "Passed":
                    cell.fill = FILL_PASSED
                    cell.font = Font(name="Outfit", color="155724", bold=True)
                elif status_val == "Failed":
                    cell.fill = FILL_FAILED
                    cell.font = Font(name="Outfit", color="721C24", bold=True)
                elif status_val == "Blocked":
                    cell.fill = FILL_BLOCKED
                    cell.font = Font(name="Outfit", color="856404", bold=True)
                elif status_val == "Skipped":
                    cell.fill = FILL_SKIPPED
                    cell.font = Font(name="Outfit", color="383D41", bold=True)
                    
    # Adjust widths automatically
    apply_auto_width(ws_det)
    
    # Save Workbook
    excel_path = "Test_Results/Excel/appium-test-report.xlsx"
    wb.save(excel_path)
    print(f"Excel test report saved successfully at: {excel_path}")
    
    # 3. Generate Markdown Summary
    md_path = "Test_Results/Markdown/appium-test-summary.md"
    md_content = f"""### Appium Mobile E2E Test Suite Summary
| Metric | Value |
| --- | --- |
| **Total Test Cases** | {total} |
| **Passed Tests** | {passed} |
| **Failed Tests** | {failed} |
| **Skipped Tests** | {skipped} |
| **Blocked Tests** | {blocked} |
| **Pass Percentage** | **{pass_percent}%** |
| **Execution Duration** | **{duration_sec:.2f} seconds** |
| **Execution Status** | **{"PASS" if pass_percent >= 95.0 else "FAIL"}** |

*Report generated automatically after Appium test execution suite finished.*
"""
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md_content)
    print(f"Markdown test report saved successfully at: {md_path}")
